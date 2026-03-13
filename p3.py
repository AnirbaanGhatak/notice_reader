import pyautogui
import pyperclip
import pandas as pd
from google import genai
import openpyxl
import time
import os
import sys
import json
from datetime import datetime, timedelta
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1: GLOBAL CONFIGURATION  ← Edit everything here before first run
# ─────────────────────────────────────────────────────────────────────────────

CLIENT_MASTER_FILE    = "winman_master.xlsx"  # Input  — client list
NOTICES_MASTER_FILE   = "gst_notices.xlsx"    # Output — notices log
PASSWORD_ERRORS_FILE  = "Password_Errors.xlsx"       # Clients with bad passwords

MAX_CLIENTS_PER_RUN   = 60    # Daily cap to avoid IP bans
DAYS_BETWEEN_CHECKS   = 5     # Skip clients checked within this many days
LOGIN_WAIT_SECONDS    = 40    # Max wait after CAPTCHA submit (polling cuts this short)
CAPTCHA_TEMP_IMAGE    = "temp_captcha.png"           # Auto-deleted after use

# ── Password error detection (image-based, no Gemini needed) ─────────────────
# How to get the reference image:
#   1. Manually log in with a wrong password so the pink banner appears.
#   2. Win+Shift+S → crop tightly around just the pink error banner text.
#   3. Save as "login_error_banner.png" next to this script.
IMG_LOGIN_ERROR       = "login_error_banner.png"

# How often (seconds) to check for the error banner during the login wait.
# 2s is a good balance between speed and CPU usage.
ERROR_POLL_INTERVAL   = 2

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2: SCREEN COORDINATES  ← Your calibrated values
# ─────────────────────────────────────────────────────────────────────────────

COORD_SEARCH_BOX          = (442, 139)   # Winman client search input
COORD_SEARCH_RESULT_FIRST = (503, 212)   # First autocomplete result in dropdown

# ── Ribbon → Services → View Notices  (3-step hover menu) ────────────────────
COORD_RIBBON_BTN          = (320, 32)    # Ribbon button — coordinate-based

# Steps 2 & 3: IMAGE-BASED
IMG_SERVICES              = "service.jpg"     # Cropped screenshot of Services menu row
IMG_VIEW_NOTICES          = "notices.jpg"     # Cropped screenshot of View Notices row

# Image match confidence. Lower to 0.7 if not found; raise to 0.9 if wrong match.
IMG_CONFIDENCE            = 0.8

# Pause after hovering Services before searching for the side submenu.
HOVER_PAUSE_SECONDS       = 0.8
# ─────────────────────────────────────────────────────────────────────────────

CAPTCHA_REGION            = (450, 596, 179, 48)  # (x, y, width, height)
COORD_CAPTCHA_INPUT       = (463, 468)   # CAPTCHA text input box
COORD_SAFE_ZONE           = (3, 249)     # Neutral click to focus the notices page

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3: PYAUTOGUI SAFETY
# ─────────────────────────────────────────────────────────────────────────────

# Move mouse to TOP-LEFT corner at any time to instantly abort the bot.
pyautogui.FAILSAFE = True
pyautogui.PAUSE    = 0.5   # Small pause between every GUI action

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4: GEMINI CLIENT INIT
# ─────────────────────────────────────────────────────────────────────────────

def init_gemini() -> genai.Client:
    """
    Creates and returns a Gemini client. Reads the API key from the
    GOOGLE_API_KEY environment variable automatically — no hardcoded key needed.
    A single client is reused for the entire run.
    """
    client = genai.Client()
    print("✅ [INIT] Gemini client initialised.")
    return client


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5: CAPTCHA SOLVING
# ─────────────────────────────────────────────────────────────────────────────

def solve_captcha(client: genai.Client) -> str | None:
    """
    1. Screenshots the CAPTCHA region to a temp file.
    2. Uploads the image to the Gemini Files API.
    3. Asks Gemini to return ONLY the alphanumeric characters.
    4. Deletes the remote file and local temp file.
    5. Returns the cleaned CAPTCHA string, or None on any failure.

    FAILSAFE: Any API or IO error returns None. The finally block guarantees
    the local temp file is always deleted even on exceptions.
    """
    print("📸 [CAPTCHA] Taking screenshot of CAPTCHA region...")
    captcha_path = CAPTCHA_TEMP_IMAGE

    try:
        captcha_img = pyautogui.screenshot(region=CAPTCHA_REGION)
        captcha_img.save(captcha_path)

        print("🧠 [CAPTCHA] Uploading to Gemini Files API...")
        img_file = client.files.upload(file=captcha_path)

        prompt = (
            "Read the text in this CAPTCHA image. "
            "Respond ONLY with the exact letters and numbers you see. "
            "No spaces, no punctuation, no extra words, no explanation."
        )

        response = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=[img_file, prompt]
        )

        client.files.delete(name=img_file.name)

        captcha_text = response.text.strip().replace(" ", "")
        print(f"✅ [CAPTCHA] Gemini read: '{captcha_text}'")
        return captcha_text

    except Exception as e:
        print(f"❌ [CAPTCHA ERROR] {e}")
        return None

    finally:
        if os.path.exists(captcha_path):
            os.remove(captcha_path)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6: GEMINI DATA CLEANING
# ─────────────────────────────────────────────────────────────────────────────

# The exact column schema written to GST_Notices_Master.xlsx.
# Company Name, GSTIN, Extraction Date are injected by us — not from the portal.
# The portal columns (first 5) are what Gemini maps from the raw clipboard text.
EXPECTED_COLUMNS = [
    "Notice/Demand Order ID",   # Unique ID — used for de-duplication
    "Type",                     # Notice form code e.g. ASMT-10, DRC-01
    "Notice/Order Description", # Subject text of the notice
    "Date of Issuance",         # When the notice was issued
    "Due Date",                 # Response deadline
    "Company Name",             # Injected by bot — not from portal
    "GSTIN",                    # Injected by bot — not from portal
    "Extraction Date",          # Injected by bot — date this run occurred
]

def clean_notices_with_gemini(
    client: genai.Client,
    raw_clipboard: str,        # Raw text copied from the notices page
    company_name: str,
    gstin: str,
) -> pd.DataFrame | None:
    """
    Sends the raw clipboard text to Gemini and asks it to extract and clean
    the notices table into our EXPECTED_COLUMNS schema.

    WHY RAW STRING INSTEAD OF pd.read_clipboard():
    The clipboard content from an embedded browser is messy HTML-derived text,
    not clean TSV. Passing the raw string directly to Gemini lets it handle all
    the parsing, column mapping, and date standardisation in one step — far more
    robust than trying to coerce it through pandas first.

    Gemini is asked to:
      ✓ Extract only the notices table rows, ignoring all surrounding text.
      ✓ Map columns to our EXPECTED_COLUMNS schema.
      ✓ Standardise all dates to YYYY-MM-DD.
      ✓ Strip whitespace from all values.
      ✓ Return ONLY a raw JSON array — no markdown, no prose.

    FAILSAFE: If Gemini returns unparseable JSON or fails entirely, we return
    None and the caller logs a warning — data for this client is skipped but
    the bot continues. We don't save garbage data.
    """
    if not raw_clipboard or not raw_clipboard.strip():
        return None

    print("🔍 [CLEAN] Sending raw clipboard to Gemini for parsing & cleaning...")

    today = datetime.today().strftime("%Y-%m-%d")

    prompt = f"""You are a data extraction assistant for a GST (Goods and Services Tax) notice tracker.

Below is raw text copied from the GST portal notices page. It may contain page headers,
navigation text, footers, and other noise around the actual notices table.

Your task:
1. Find and extract ONLY the rows that represent GST notices/demand orders.
2. Map each row to this EXACT column schema:
   - "Notice/Demand Order ID"   → the unique notice or demand order reference number
   - "Type"                     → the notice form code (e.g. ASMT-10, DRC-01, SCN, etc.)
   - "Notice/Order Description" → the subject or description text of the notice
   - "Date of Issuance"         → date the notice was issued, formatted as YYYY-MM-DD
   - "Due Date"                 → response deadline, formatted as YYYY-MM-DD
   - "Company Name"             → always set to exactly: "{company_name}"
   - "GSTIN"                    → always set to exactly: "{gstin}"
   - "Extraction Date"          → always set to exactly: "{today}"

Cleaning rules:
- Strip leading/trailing whitespace from all string values.
- Standardise ALL dates to YYYY-MM-DD format. If a date cannot be parsed, use null.
- If there are no notices at all (e.g. "No records found"), return an empty array: []
- Do NOT invent data. If a field has no match, use null.
- Ignore all text that is not part of the notices table (headers, nav, footers, etc.).

Output format:
Return ONLY a valid JSON array starting with [ and ending with ].
No markdown, no code fences, no explanation whatsoever.

Raw clipboard text:
{raw_clipboard}"""

    try:
        response = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=prompt
        )

        cleaned_text = response.text.strip()

        # Strip markdown code fences if Gemini adds them despite instructions
        if cleaned_text.startswith("```"):
            cleaned_text = "\n".join(
                line for line in cleaned_text.splitlines()
                if not line.strip().startswith("```")
            ).strip()

        cleaned_records = json.loads(cleaned_text)

        # Empty array means no notices — valid result, not an error
        if not cleaned_records:
            print("  ℹ️  [CLEAN] Gemini found no notice rows in the clipboard content.")
            return pd.DataFrame(columns=EXPECTED_COLUMNS)

        cleaned_df = pd.DataFrame(cleaned_records)

        # Guarantee every expected column exists
        for col in EXPECTED_COLUMNS:
            if col not in cleaned_df.columns:
                cleaned_df[col] = None

        cleaned_df = cleaned_df[EXPECTED_COLUMNS]
        print(f"✅ [CLEAN] Gemini extracted {len(cleaned_df)} notice row(s).")
        return cleaned_df

    except json.JSONDecodeError as e:
        print(f"⚠️  [CLEAN] Gemini returned invalid JSON ({e}) — skipping this client.")
        return None

    except Exception as e:
        print(f"⚠️  [CLEAN] Gemini cleaning failed ({e}) — skipping this client.")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7: PASSWORD ERROR HANDLING
# ─────────────────────────────────────────────────────────────────────────────

class PasswordError(Exception):
    """
    Raised when the bot detects the GST portal's 'Invalid Username or Password'
    error banner during the login wait. Treated differently from other errors:
      - Logged to Password_Errors.xlsx so you know exactly which clients to fix.
      - Last Checked Date is NOT updated, so the client is retried next run
        once the password has been corrected in Winman.
    """
    pass


def log_password_error(company_name: str, gstin: str) -> None:
    """
    Appends one row to Password_Errors.xlsx with the client details and
    the timestamp of when the error was detected.

    The file accumulates across all runs — each entry is a separate incident
    so you have a full history of when each password started failing.
    Open it at any time to see which clients need their passwords fixed in Winman.
    """
    now = datetime.now()
    new_row = pd.DataFrame([{
        "Company Name" : company_name,
        "GSTIN"        : gstin,
        "Error Date"   : now.strftime("%Y-%m-%d"),
        "Error Time"   : now.strftime("%H:%M:%S"),
        "Notes"        : "Incorrect password — update in Winman and re-run",
    }])

    if Path(PASSWORD_ERRORS_FILE).exists():
        existing = pd.read_excel(PASSWORD_ERRORS_FILE, dtype=str)
        combined = pd.concat([existing, new_row], ignore_index=True)
    else:
        combined = new_row

    with pd.ExcelWriter(PASSWORD_ERRORS_FILE, engine="openpyxl", mode="w") as writer:
        combined.to_excel(writer, index=False, sheet_name="Password Errors")
        ws = writer.sheets["Password Errors"]
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value), default=10
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 45)

    print(f"  📋 [PASSWORD LOG] Logged → {PASSWORD_ERRORS_FILE}: {company_name} ({gstin})")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8: UI NAVIGATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _find_on_screen(image_path: str, label: str) -> tuple[int, int] | None:
    """
    Locates a UI element on screen by image and returns its centre (x, y).
    Returns None if not found — caller decides how to handle it.
    """
    try:
        location = pyautogui.locateCenterOnScreen(
            image_path,
            confidence=IMG_CONFIDENCE
        )
        if location:
            print(f"  [IMG] Found '{label}' at ({location.x}, {location.y}).")
            return (location.x, location.y)
        print(f"  [IMG] '{label}' not found on screen.")
        return None
    except pyautogui.ImageNotFoundException:
        print(f"  [IMG] '{label}' not found on screen (ImageNotFoundException).")
        return None
    except Exception as e:
        print(f"  [IMG ERROR] Searching for '{label}': {e}")
        return None


def navigate_to_client(company_name: str) -> None:
    """
    Full navigation sequence:
    A. Type company name in search → select first result.
    B. Click ribbon button → first dropdown (coordinate-based).
    C. Image-search for 'Services' → moveTo() to hover (opens side submenu).
    D. Image-search for 'View Notices' → click (triggers Winman login sequence).

    FAILSAFE: RuntimeError raised if either image is not found — caught by
    process_client which resets the UI and moves to the next client.
    """
    print(f"  [NAV] Searching Winman for: {company_name}")

    # ── A: Select the client ──────────────────────────────────────────────────
    pyautogui.click(*COORD_SEARCH_BOX)
    time.sleep(0.5)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("delete")
    pyautogui.typewrite(company_name, interval=0.05)
    time.sleep(1.5)
    pyautogui.click(*COORD_SEARCH_RESULT_FIRST)
    time.sleep(0.8)
    print("  [NAV] Client selected.")

    # ── B: Open first dropdown ────────────────────────────────────────────────
    pyautogui.click(*COORD_RIBBON_BTN)
    time.sleep(0.6)
    print("  [NAV] Ribbon dropdown opened.")

    # ── C: Hover over Services to open side submenu ───────────────────────────
    # moveTo not click — clicking would close the submenu before we reach it
    services_pos = _find_on_screen(IMG_SERVICES, "Services")
    if not services_pos:
        raise RuntimeError(
            f"'Services' image ({IMG_SERVICES}) not found. "
            "Check the file exists and IMG_CONFIDENCE is not too high."
        )
    pyautogui.moveTo(*services_pos, duration=0.3)
    time.sleep(HOVER_PAUSE_SECONDS)
    print("  [NAV] Hovering on 'Services' — side submenu should be visible.")

    # ── D: Click View Notices ─────────────────────────────────────────────────
    view_notices_pos = _find_on_screen(IMG_VIEW_NOTICES, "View Notices")
    if not view_notices_pos:
        raise RuntimeError(
            f"'View Notices' image ({IMG_VIEW_NOTICES}) not found. "
            "Submenu may not have opened — try increasing HOVER_PAUSE_SECONDS."
        )
    pyautogui.click(*view_notices_pos)
    time.sleep(2.0)
    print("  [NAV] 'View Notices' clicked — GST Portal login screen loading.")


def perform_login(captcha_text: str) -> None:
    """
    Types the CAPTCHA and submits the form, then polls for the pink error
    banner every ERROR_POLL_INTERVAL seconds for up to LOGIN_WAIT_SECONDS.

    WHY POLLING INSTEAD OF A FLAT SLEEP:
    A wrong password shows the pink error banner typically within 2–4 seconds.
    Polling detects it immediately, raises PasswordError, and saves the remaining
    ~36 seconds of wait time per failed client.

    OUTCOMES:
      Error banner detected  → raises PasswordError (caught in process_client)
      Full wait with no banner → returns normally → proceed to extraction
      IMG_LOGIN_ERROR missing  → falls back to flat sleep (bot keeps running)

    FAILSAFE: If the image file is missing we print a tip and do the flat wait
    so the bot is never blocked by a missing reference image.
    """
    print(f"  [LOGIN] Entering CAPTCHA: '{captcha_text}'")
    pyautogui.click(*COORD_CAPTCHA_INPUT)
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("delete")
    pyautogui.typewrite(captcha_text, interval=0.05)
    pyautogui.press("enter")

    # ── Check reference image exists ──────────────────────────────────────────
    if not Path(IMG_LOGIN_ERROR).exists():
        print(f"  ⚠️  [LOGIN] '{IMG_LOGIN_ERROR}' not found — flat {LOGIN_WAIT_SECONDS}s wait.")
        print(f"  💡 Tip: snip the pink error banner and save as '{IMG_LOGIN_ERROR}' for fast detection.")
        time.sleep(LOGIN_WAIT_SECONDS)
        return

    # ── Polling loop ──────────────────────────────────────────────────────────
    print(f"  [LOGIN] Polling for error banner every {ERROR_POLL_INTERVAL}s (max {LOGIN_WAIT_SECONDS}s)...")
    elapsed = 0
    while elapsed < LOGIN_WAIT_SECONDS:
        time.sleep(ERROR_POLL_INTERVAL)
        elapsed += ERROR_POLL_INTERVAL

        try:
            found = pyautogui.locateOnScreen(IMG_LOGIN_ERROR, confidence=IMG_CONFIDENCE)
        except pyautogui.ImageNotFoundException:
            found = None
        except Exception as e:
            print(f"  ⚠️  [LOGIN] Banner scan error ({e}) — continuing wait.")
            found = None

        if found:
            # Fail fast — no point waiting the remaining time
            raise PasswordError(
                "GST portal shows 'Invalid Username or Password' banner. "
                "Please update this client's password in Winman."
            )

        print(f"  [LOGIN] {elapsed}s — no error banner, still waiting...")

    print(f"  ✅ [LOGIN] No error banner after {LOGIN_WAIT_SECONDS}s — login presumed successful.")


def reset_to_main_screen() -> None:
    """
    POST-FLOW RESET — ALWAYS called after each client, success or failure.
    Presses Esc 3× to close the portal window and return to Winman's main screen.
    This guarantees every new client iteration starts from a clean UI state.
    """
    print("  [RESET] Pressing Esc × 3...")
    for _ in range(3):
        pyautogui.press("esc")
        time.sleep(0.5)
    time.sleep(1.0)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9: DATA EXTRACTION (Clipboard Method)
# ─────────────────────────────────────────────────────────────────────────────

def extract_raw_clipboard() -> str | None:
    """
    Clicks the safe-zone to focus the embedded browser, then uses Ctrl+A →
    Ctrl+C to copy everything on the notices page to the clipboard.

    Returns the raw clipboard TEXT STRING (not a DataFrame). The raw string
    is passed directly to clean_notices_with_gemini() which handles all
    parsing — this avoids pd.read_clipboard()'s brittle TSV assumptions on
    messy browser-copied content.

    FAILSAFE: Empty clipboard returns None — caller logs a warning and
    continues to update the date and move to the next client.
    """
    print("  [EXTRACT] Focusing notices page and copying content...")
    pyautogui.click(*COORD_SAFE_ZONE)
    time.sleep(0.5)

    # Clear clipboard first so we can detect a failed copy
    pyperclip.copy("")
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "c")
    time.sleep(1.5)   # Give the OS time to complete the copy

    content = pyperclip.paste()
    if not content or not content.strip():
        print("  [EXTRACT ERROR] Clipboard is empty — notices page may not have loaded.")
        return None

    print(f"  [EXTRACT] Got {len(content)} characters of raw content from clipboard.")
    return content


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10: EXCEL READ / WRITE (pandas + openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

def load_client_master() -> pd.DataFrame:
    """Loads Winman_Client_Master.xlsx. Expected columns: Company Name, GSTIN, Last Checked Date."""
    if not Path(CLIENT_MASTER_FILE).exists():
        print(f"❌ [ERROR] Client master not found: {CLIENT_MASTER_FILE}")
        sys.exit(1)
    df = pd.read_excel(CLIENT_MASTER_FILE, dtype=str)
    df.columns = df.columns.str.strip()
    df["Last Checked Date"] = pd.to_datetime(df["Last Checked Date"], errors="coerce")
    print(f"✅ [LOAD] {len(df)} clients loaded.")
    return df


def filter_eligible_clients(df: pd.DataFrame) -> pd.DataFrame:
    """
    Returns clients not checked within DAYS_BETWEEN_CHECKS days, capped at
    MAX_CLIENTS_PER_RUN. Never-checked clients (NaT date) are always included.
    """
    cutoff = datetime.today() - timedelta(days=DAYS_BETWEEN_CHECKS)
    mask   = df["Last Checked Date"].isna() | (df["Last Checked Date"] < cutoff)
    result = df[mask].copy()
    print(f"📋 [FILTER] {len(result)} eligible clients (daily cap: {MAX_CLIENTS_PER_RUN}).")
    return result.head(MAX_CLIENTS_PER_RUN)


def update_last_checked_date(company_name: str) -> None:
    """
    Updates 'Last Checked Date' for the given company to today.
    pandas finds the row; openpyxl writes only that one cell — preserving
    any manual formatting or formulas in the rest of the workbook.
    """
    today_str = datetime.today().strftime("%Y-%m-%d")

    df = pd.read_excel(CLIENT_MASTER_FILE, dtype=str)
    df.columns = df.columns.str.strip()
    match = df["Company Name"].str.strip() == company_name.strip()

    if not match.any():
        print(f"  ⚠️  [UPDATE] '{company_name}' not found in client master.")
        return

    # pandas 0-based index → openpyxl row (+2: 1-based + header row)
    excel_row = df[match].index[0] + 2

    wb = openpyxl.load_workbook(CLIENT_MASTER_FILE)
    ws = wb.active

    header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
    date_col   = header_map.get("Last Checked Date")

    if not date_col:
        print("  ⚠️  [UPDATE] 'Last Checked Date' column not found in workbook.")
        wb.close()
        return

    ws.cell(row=excel_row, column=date_col).value = today_str
    wb.save(CLIENT_MASTER_FILE)
    wb.close()
    print(f"  ✅ [UPDATE] Last Checked Date → {today_str} for '{company_name}'.")


def append_to_notices_master(cleaned_df: pd.DataFrame) -> None:
    """
    Merges new cleaned notices into GST_Notices_Master.xlsx:
      1. Load existing rows (create fresh if file doesn't exist).
      2. Concatenate new rows.
      3. De-duplicate on 'Notice/Demand Order ID' — keep LAST occurrence.
      4. Enforce EXPECTED_COLUMNS order.
      5. Write back with auto-fitted column widths.
    """
    if cleaned_df is None or cleaned_df.empty:
        print("  ℹ️  [SAVE] No new notices to append.")
        return

    if Path(NOTICES_MASTER_FILE).exists():
        existing_df = pd.read_excel(NOTICES_MASTER_FILE, dtype=str)
        print(f"  [SAVE] Existing master: {len(existing_df)} rows.")
    else:
        existing_df = pd.DataFrame(columns=EXPECTED_COLUMNS)
        print("  [SAVE] Creating new notices master file.")

    combined_df = pd.concat([existing_df, cleaned_df], ignore_index=True)

    # De-duplicate on the unique notice ID — keep LAST (most recently extracted)
    dedup_col = "Notice/Demand Order ID"
    if dedup_col in combined_df.columns:
        before = len(combined_df)
        combined_df.drop_duplicates(subset=[dedup_col], keep="last", inplace=True)
        removed = before - len(combined_df)
        if removed:
            print(f"  [DEDUP] Removed {removed} duplicate notice(s).")

    # Enforce column order — fill any gaps with None
    for col in EXPECTED_COLUMNS:
        if col not in combined_df.columns:
            combined_df[col] = None
    combined_df = combined_df[EXPECTED_COLUMNS]

    with pd.ExcelWriter(NOTICES_MASTER_FILE, engine="openpyxl", mode="w") as writer:
        combined_df.to_excel(writer, index=False, sheet_name="Notices")
        ws = writer.sheets["Notices"]
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value), default=10
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    print(f"  ✅ [SAVE] {NOTICES_MASTER_FILE} updated — {len(combined_df)} total rows.")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 11: CORE PER-CLIENT LOOP
# ─────────────────────────────────────────────────────────────────────────────

def process_client(client: genai.Client, company_name: str, gstin: str) -> bool:
    """
    Full automation sequence for one client. Returns True on success.

    FAILSAFE TABLE:
    ┌─────────────────────┬──────────────────────────────────────────────────────┐
    │ Step                │ On Failure                                           │
    ├─────────────────────┼──────────────────────────────────────────────────────┤
    │ Navigation          │ reset() → return False                               │
    │ CAPTCHA             │ retry once → if still None → reset() → return False  │
    │ Login (submit)      │ generic error → reset() → return False               │
    │ Login (polling)     │ PasswordError → log_password_error() + reset() +     │
    │                     │   return False (Last Checked Date NOT updated)        │
    │ Extract clipboard   │ log warning → continue (still update Last Checked)   │
    │ Gemini clean        │ returns None → no data written, date still updated   │
    │ Save Excel          │ log error → still reset UI                           │
    │ Reset UI            │ ALWAYS runs regardless of any outcome above          │
    └─────────────────────┴──────────────────────────────────────────────────────┘

    NOTE on Password errors: Last Checked Date is intentionally NOT updated so
    the client stays eligible for the next run — ready to succeed once the
    password is fixed in Winman.
    """
    print(f"\n{'─'*62}")
    print(f"  🏢 {company_name}  |  GSTIN: {gstin}")
    print(f"{'─'*62}")

    # ── Step 1: Navigate ──────────────────────────────────────────────────────
    try:
        navigate_to_client(company_name)
    except Exception as e:
        print(f"  ❌ [NAV ERROR] {e}")
        reset_to_main_screen()
        return False

    # ── Step 2: Solve CAPTCHA (one automatic retry) ───────────────────────────
    try:
        captcha_text = solve_captcha(client)
        if not captcha_text:
            print("  ⚠️  [CAPTCHA] Attempt 1 empty — retrying in 2s...")
            time.sleep(2)
            captcha_text = solve_captcha(client)
        if not captcha_text:
            print("  ❌ [CAPTCHA] Both attempts failed — skipping client.")
            reset_to_main_screen()
            return False
    except Exception as e:
        print(f"  ❌ [CAPTCHA ERROR] {e}")
        reset_to_main_screen()
        return False

    # ── Step 3: Login with continuous error-banner polling ────────────────────
    # perform_login() polls for the pink error banner every ERROR_POLL_INTERVAL
    # seconds. Raises PasswordError on detection — saves the remaining wait time.
    try:
        perform_login(captcha_text)

    except PasswordError as e:
        # Wrong password — log it and skip WITHOUT updating Last Checked Date
        print(f"  🔑 [PASSWORD ERROR] {e}")
        log_password_error(company_name, gstin)
        reset_to_main_screen()
        return False

    except Exception as e:
        print(f"  ❌ [LOGIN ERROR] {e}")
        reset_to_main_screen()
        return False

    # ── Step 4: Copy raw clipboard content from notices page ──────────────────
    raw_clipboard = None
    try:
        raw_clipboard = extract_raw_clipboard()
        if raw_clipboard is None:
            print("  ⚠️  [EXTRACT] Nothing copied — marking as checked and continuing.")
    except Exception as e:
        print(f"  ❌ [EXTRACT ERROR] {e}")

    # ── Step 5: Gemini parses + cleans the raw clipboard text ─────────────────
    cleaned_df = None
    if raw_clipboard:
        try:
            cleaned_df = clean_notices_with_gemini(
                client, raw_clipboard, company_name, gstin
            )
        except Exception as e:
            print(f"  ❌ [CLEAN ERROR] {e} — no data written for this client.")

    # ── Step 6: Save to Excel ─────────────────────────────────────────────────
    try:
        append_to_notices_master(cleaned_df)
        update_last_checked_date(company_name)
    except Exception as e:
        print(f"  ❌ [SAVE ERROR] {e}")

    # ── Step 7: Reset UI — ALWAYS runs ───────────────────────────────────────
    reset_to_main_screen()
    return True


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 12: MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  🤖 WINMAN GST NOTICE CHECKER — STARTING")
    print(f"  Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)
    print("\n⚠️  FAILSAFE: Move mouse to TOP-LEFT corner to abort instantly.")
    print("⏳ Focus the Winman window now — starting in 5 seconds...\n")
    time.sleep(5)

    client      = init_gemini()
    client_df   = load_client_master()
    eligible_df = filter_eligible_clients(client_df)

    if eligible_df.empty:
        print("✅ [DONE] No clients need checking today. Exiting.")
        return

    total, succeeded, failed = len(eligible_df), 0, 0

    for idx, row in eligible_df.iterrows():
        company_name = str(row.get("Company Name", "")).strip()
        gstin        = str(row.get("GSTIN",        "")).strip()

        if not company_name or not gstin:
            print(f"⚠️  [SKIP] Row {idx}: Missing Company Name or GSTIN.")
            failed += 1
            continue

        success = process_client(client, company_name, gstin)
        if success:
            succeeded += 1
            print(f"  ✅ ({succeeded + failed}/{total}) Done: {company_name}")
        else:
            failed += 1
            print(f"  ❌ ({succeeded + failed}/{total}) Failed: {company_name}")

        time.sleep(3)   # Polite pause between clients

    # ── Run summary ───────────────────────────────────────────────────────────
    print("\n" + "=" * 62)
    print(f"  🏁 RUN COMPLETE")
    print(f"  Total    : {total}")
    print(f"  ✅ Passed : {succeeded}")
    print(f"  ❌ Failed : {failed}")

    if Path(PASSWORD_ERRORS_FILE).exists():
        err_df = pd.read_excel(PASSWORD_ERRORS_FILE, dtype=str)
        print(f"  🔑 Password errors on file : {len(err_df)} total → {PASSWORD_ERRORS_FILE}")
        print(f"     Fix these passwords in Winman, then re-run.")

    print(f"  Ended at : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)


    main()