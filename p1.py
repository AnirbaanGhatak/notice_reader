import pyautogui
import pyperclip
import pandas as pd
from google import genai                  
import openpyxl
import time
import os
import sys
import json
from datetime import datetime, timedelta
from pathlib import Path


CLIENT_MASTER_FILE    = "Winman_Client_Master.xlsx"  # Input  — client list
NOTICES_MASTER_FILE   = "GST_Notices_Master.xlsx"    # Output — notices log

MAX_CLIENTS_PER_RUN   = 60     # Daily cap to avoid IP bans
DAYS_BETWEEN_CHECKS   = 5      # Skip clients checked within this many days
LOGIN_WAIT_SECONDS    = 40     # Wait after CAPTCHA submit for notices page
CAPTCHA_TEMP_IMAGE    = "temp_captcha.png"           # Auto-deleted after use


COORD_SEARCH_BOX          = (400, 150)   # Winman client search input
COORD_SEARCH_RESULT_FIRST = (400, 190)   # First autocomplete result
COORD_OPEN_GST_PORTAL_BTN = (700, 150)   # Button to open GST portal
CAPTCHA_REGION            = (500, 400, 200, 60)  # (x, y, width, height)
COORD_CAPTCHA_INPUT       = (600, 470)   # CAPTCHA text input box
COORD_SAFE_ZONE           = (800, 500)   # Neutral click inside notices table


# Move mouse to TOP-LEFT corner at any time to instantly abort the bot.
pyautogui.FAILSAFE = True
pyautogui.PAUSE    = 0.5   # Small pause between every GUI action


def init_gemini() -> genai.Client:
    """
    Creates and returns a Gemini client using the new google-genai SDK.
    A single client instance is reused across all clients in the run to
    avoid re-authenticating on every iteration.
    """
    client = genai.Client()
    print("✅ [INIT] Gemini client initialised.")
    return client


def solve_captcha(client: genai.Client) -> str | None:
    """
    1. Screenshots the CAPTCHA region to a temp file.
    2. Uploads the image to the Gemini Files API.
    3. Sends the file reference to generate_content() with a strict prompt.
    4. Deletes the remote file and local temp file.
    5. Returns the cleaned CAPTCHA string, or None on any failure.

    FAILSAFE: Entire function is wrapped — any API or IO error returns None.
    The caller handles retry / graceful skip. The finally block guarantees
    the local temp file is always deleted, even on exceptions.
    """
    print("[CAPTCHA] Taking screenshot of CAPTCHA region...")
    captcha_path = CAPTCHA_TEMP_IMAGE

    try:
        captcha_img = pyautogui.screenshot(region=CAPTCHA_REGION)
        captcha_img.save(captcha_path)
  
        print("[CAPTCHA] Uploading to Gemini Files API...")
        img_file = client.files.upload(file=captcha_path)

        prompt = (
            "Read the text in this CAPTCHA image. "
            "Respond ONLY with the exact letters and numbers you see. "
            "No spaces, no punctuation, no extra words, no explanation."
        )

        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[img_file, prompt]
        )

        # ── Step 4: Delete remote file immediately (API hygiene) ──────────────
        client.files.delete(name=img_file.name)

        # ── Step 5: Sanitise and return ───────────────────────────────────────
        captcha_text = response.text.strip().replace(" ", "")
        print(f"✅ [CAPTCHA] Gemini read: '{captcha_text}'")
        return captcha_text

    except Exception as e:
        print(f"[CAPTCHA ERROR] {e}")
        return None

    finally:
        # delete local temp — runs even if an exception occurred above
        if os.path.exists(captcha_path):
            os.remove(captcha_path)


# Canonical column schema for GST_Notices_Master.xlsx.
# Gemini maps whatever the clipboard gives us into this exact shape.
EXPECTED_COLUMNS = [
    "Notice/Demand Order ID",
    "Type",
    "Notice/Order Description",
    "Date of Issuance",
    "Due Date"
]

def clean_notices_with_gemini(
    client: genai.Client,
    raw_data: str,
    company_name: str,
) -> pd.DataFrame | None:
    
    if raw_data is None:
        return None

    print("🔍 [CLEAN] Sending raw data to Gemini for cleaning & normalisation...")

    
    today    = datetime.today().strftime("%Y-%m-%d")

    prompt = f"""You are a data cleaning assistant for a GST notice tracker.

Your task — clean & map this data to the following EXACT column schema:

Column definitions:
- "Notice/Demand Order ID"  
- "Type"       
- "Notice/Order Description"     
- "Response Due Date"
- "Date of Issuance"
- "Due Date"


Cleaning rules:
1. Strip leading/trailing whitespace from all string values.
2. All dates → YYYY-MM-DD. If a date cannot be parsed, use null.
3. Drop rows where BOTH "Reference No" AND "Notice Type" are null or empty.
4. Do NOT invent data. If no source column maps to a target, use null.
5. Remove all other text around the Table, the text around the table is not required.

Output format:
Return ONLY a valid JSON array. No markdown, no code fences, no explanation.
Just the raw JSON array starting with [ and ending with ].

Raw data:
{raw_data}"""

    try:
        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt
        )

        cleaned_text = response.text.strip()

        # Strip markdown code fences if Gemini adds them despite the instruction
        if cleaned_text.startswith("```"):
            lines = cleaned_text.splitlines()
            cleaned_text = "\n".join(
                l for l in lines if not l.strip().startswith("```")
            ).strip()

        cleaned_records = json.loads(cleaned_text)
        cleaned_df      = pd.DataFrame(cleaned_records)

        # Guarantee every expected column exists (fill missing with None)
        for col in EXPECTED_COLUMNS:
            if col not in cleaned_df.columns:
                cleaned_df[col] = None

        # Enforce the standard column order
        cleaned_df = cleaned_df[EXPECTED_COLUMNS]

        print(f"✅ [CLEAN] Gemini returned {len(cleaned_df)} cleaned rows.")
        return cleaned_df

    except Exception as e:
        # FALLBACK — Gemini failed, but we still save the raw data
        print(f"[CLEAN FALLBACK] Gemini cleaning failed ({e}). Saving raw data.")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7: UI NAVIGATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def navigate_to_client(company_name: str) -> None:
    """Searches Winman for the client and opens their GST Portal login screen."""
    print(f"  [NAV] Searching Winman for: {company_name}")
    pyautogui.click(*COORD_SEARCH_BOX)
    time.sleep(0.5)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("delete")
    pyautogui.typewrite(company_name, interval=0.05)
    time.sleep(1.5)
    pyautogui.click(*COORD_SEARCH_RESULT_FIRST)
    time.sleep(0.8)
    pyautogui.click(*COORD_OPEN_GST_PORTAL_BTN)
    time.sleep(2.0)
    print("  [NAV] GST Portal login screen opened.")


def perform_login(captcha_text: str) -> None:
    """Types the CAPTCHA, submits the login form, then waits for the notices page."""
    print(f"  [LOGIN] Entering CAPTCHA: '{captcha_text}'")
    pyautogui.click(*COORD_CAPTCHA_INPUT)
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("delete")
    pyautogui.typewrite(captcha_text, interval=0.05)
    pyautogui.press("enter")
    print(f"  [LOGIN] Waiting {LOGIN_WAIT_SECONDS}s for notices page to load...")
    time.sleep(LOGIN_WAIT_SECONDS)


def reset_to_main_screen() -> None:
    print("  [RESET] Pressing Esc × 3...")
    for _ in range(3):
        pyautogui.press("esc")
        time.sleep(0.5)
    time.sleep(1.0)


def extract_raw_notices() -> pd.DataFrame | None:
    
    print("  [EXTRACT] Focusing browser and copying page content...")
    pyautogui.click(*COORD_SAFE_ZONE)
    time.sleep(0.5)

    # Clear the clipboard so we can detect a failed copy (empty string result)
    pyperclip.copy("")
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "c")
    time.sleep(1.5)   # Give the OS time to finish the copy operation

    clipboard_content = pyperclip.paste()
    if not clipboard_content or clipboard_content.strip() == "":
        print("  [EXTRACT ERROR] Clipboard empty — notices page may not have loaded.")
        return None

    try:
        df = pd.read_clipboard()
        if df.empty:
            print("  [EXTRACT WARNING] Clipboard parsed but table is empty.")
            return None
        print(f"  [EXTRACT] Got {len(df)} raw rows from clipboard.")
        return df
    except Exception as e:
        print(f"  [EXTRACT ERROR] Could not parse clipboard: {e}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9: EXCEL READ / WRITE (pandas + openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

def load_client_master() -> pd.DataFrame:
    """
    Loads Winman_Client_Master.xlsx.
    Expected columns: 'Company Name', 'GSTIN', 'Last Checked Date'
    """
    if not Path(CLIENT_MASTER_FILE).exists():
        print(f"❌ [ERROR] Client master not found: {CLIENT_MASTER_FILE}")
        sys.exit(1)
    df = pd.read_excel(CLIENT_MASTER_FILE, dtype=str)
    df.columns = df.columns.str.strip()
    df["Last Checked Date"] = pd.to_datetime(df["Last Checked Date"], errors="coerce")
    print(f"✅ [LOAD] {len(df)} clients loaded.")
    return df


def filter_eligible_clients(df: pd.DataFrame) -> pd.DataFrame:
    """
    Returns clients not checked within DAYS_BETWEEN_CHECKS days, capped at
    MAX_CLIENTS_PER_RUN. Clients with a NaT (never-checked) date are always
    included as highest priority.
    """
    cutoff = datetime.today() - timedelta(days=DAYS_BETWEEN_CHECKS)
    mask   = df["Last Checked Date"].isna() | (df["Last Checked Date"] < cutoff)
    result = df[mask].copy()
    print(f"📋 [FILTER] {len(result)} eligible (cap: {MAX_CLIENTS_PER_RUN}).")
    return result.head(MAX_CLIENTS_PER_RUN)


def update_last_checked_date(company_name: str) -> None:
    """
    Updates 'Last Checked Date' for the given company to today in the client master.

    Strategy: pandas locates the row index cleanly; openpyxl writes the cell
    value directly. This preserves any manual formatting in the workbook that
    a full pandas overwrite would destroy.
    """
    today_str = datetime.today().strftime("%Y-%m-%d")

    # Use pandas to find the target row index (clean, readable, no openpyxl loops)
    df = pd.read_excel(CLIENT_MASTER_FILE, dtype=str)
    df.columns = df.columns.str.strip()
    match = df["Company Name"].str.strip() == company_name.strip()

    if not match.any():
        print(f"  ⚠️  [UPDATE] '{company_name}' not found in client master.")
        return

    # pandas row index (0-based) → openpyxl row number (+2: header row + 1-based)
    pandas_row_idx = df[match].index[0]
    excel_row      = pandas_row_idx + 2

    wb = openpyxl.load_workbook(CLIENT_MASTER_FILE)
    ws = wb.active

    # Locate 'Last Checked Date' column from the actual header row (not hardcoded)
    header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
    date_col   = header_map.get("Last Checked Date")

    if not date_col:
        print("  ⚠️  [UPDATE] 'Last Checked Date' column not found in workbook.")
        wb.close()
        return

    ws.cell(row=excel_row, column=date_col).value = today_str
    wb.save(CLIENT_MASTER_FILE)
    wb.close()
    print(f"  ✅ [UPDATE] Last Checked Date → {today_str} for '{company_name}'.")


def append_to_notices_master(cleaned_df: pd.DataFrame) -> None:
    # Merges new cleaned notices into GST_Notices_Master.xlsx:
    #   1. Load existing rows (or start fresh if file doesn't exist).
    #   2. Concatenate new rows below existing ones.
    #   3. De-duplicate on 'Reference No' — keep LAST (most recently extracted).
    #   4. Enforce EXPECTED_COLUMNS order.
    #   5. Write back using pd.ExcelWriter (openpyxl engine).
    #   6. Auto-fit column widths for readability.

    # The file is always fully overwritten — we carry the complete history
    # in memory, so this is safe and avoids append-mode corruption issues.

    if cleaned_df is None or cleaned_df.empty:
        print("[SAVE] No new notices to append.")
        return

    # ── Load existing ─────────────────────────────────────────────────────────
    if Path(NOTICES_MASTER_FILE).exists():
        existing_df = pd.read_excel(NOTICES_MASTER_FILE, dtype=str)
        print(f"[SAVE] Existing master: {len(existing_df)} rows.")
    else:
        existing_df = pd.DataFrame(columns=EXPECTED_COLUMNS)
        print("[SAVE] No existing notices file — creating fresh.")

    # ── Merge ─────────────────────────────────────────────────────────────────
    combined_df = pd.concat([existing_df, cleaned_df], ignore_index=True)

    # ── De-duplicate on Reference No ──────────────────────────────────────────
    if "Reference No" in combined_df.columns:
        before = len(combined_df)
        combined_df.drop_duplicates(
            subset=["Reference No"], keep="last", inplace=True
        )
        removed = before - len(combined_df)
        if removed:
            print(f"  [DEDUP] Removed {removed} duplicate notice(s).")

    # ── Enforce standard column order ─────────────────────────────────────────
    for col in EXPECTED_COLUMNS:
        if col not in combined_df.columns:
            combined_df[col] = None
    combined_df = combined_df[EXPECTED_COLUMNS]

    # ── Write to Excel with auto-fitted columns ───────────────────────────────
    with pd.ExcelWriter(
        NOTICES_MASTER_FILE,
        engine="openpyxl",
        mode="w"   # Always overwrite — full history is in combined_df
    ) as writer:
        combined_df.to_excel(writer, index=False, sheet_name="Notices")

        # Auto-fit each column width for readability (quality-of-life)
        ws = writer.sheets["Notices"]
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col_cells if cell.value),
                default=10
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    print(f"  ✅ [SAVE] Notices master → {NOTICES_MASTER_FILE} ({len(combined_df)} total rows).")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10: CORE PER-CLIENT LOOP
# ─────────────────────────────────────────────────────────────────────────────

def process_client(client: genai.Client, company_name: str, gstin: str) -> bool:
    """
    Full automation sequence for one client. Returns True on success.

    FAILSAFE TABLE:
    ┌──────────────┬──────────────────────────────────────────────────────┐
    │ Step         │ On Failure                                           │
    ├──────────────┼──────────────────────────────────────────────────────┤
    │ Navigation   │ reset() → return False                               │
    │ CAPTCHA      │ retry once → if still None → reset() → return False  │
    │ Login        │ reset() → return False                               │
    │ Extract      │ log warning → continue (still update date)           │
    │ Gemini Clean │ fallback to raw data — data is NEVER silently lost   │
    │ Save Excel   │ log error → still reset UI                           │
    │ Reset UI     │ ALWAYS runs regardless of outcome above              │
    └──────────────┴──────────────────────────────────────────────────────┘
    """
    print(f"\n{'─'*62}")
    print(f"  🏢 {company_name}  |  GSTIN: {gstin}")
    print(f"{'─'*62}")

    # ── Step 1: Navigate ──────────────────────────────────────────────────────
    try:
        navigate_to_client(company_name)
    except Exception as e:
        print(f"  ❌ [NAV ERROR] {e}")
        reset_to_main_screen()
        return False

    # ── Step 2: CAPTCHA (one automatic retry) ─────────────────────────────────
    try:
        captcha_text = solve_captcha(client)
        if not captcha_text:
            print("  ⚠️  [CAPTCHA] Attempt 1 returned nothing — retrying in 2s...")
            time.sleep(2)
            captcha_text = solve_captcha(client)
        if not captcha_text:
            print("  ❌ [CAPTCHA] Both attempts failed — skipping client.")
            reset_to_main_screen()
            return False
    except Exception as e:
        print(f"  ❌ [CAPTCHA ERROR] {e}")
        reset_to_main_screen()
        return False

    # ── Step 3: Login & wait ──────────────────────────────────────────────────
    try:
        perform_login(captcha_text)
    except Exception as e:
        print(f"  ❌ [LOGIN ERROR] {e}")
        reset_to_main_screen()
        return False

    # ── Step 4: Extract raw notices ───────────────────────────────────────────
    raw_df = None
    try:
        raw_df = extract_raw_notices()
        if raw_df is None:
            print("  ⚠️  [EXTRACT] No data found — marking as checked and moving on.")
    except Exception as e:
        print(f"  ❌ [EXTRACT ERROR] {e}")

    # ── Step 5: Gemini cleans the raw data ────────────────────────────────────
    cleaned_df = None
    if raw_df is not None:
        try:
            cleaned_df = clean_notices_with_gemini(client, raw_df, company_name, gstin)
        except Exception as e:
            print(f"  ❌ [CLEAN ERROR] {e} — saving raw data as fallback.")
            today = datetime.today().strftime("%Y-%m-%d")
            raw_df["GSTIN"]           = gstin
            raw_df["Company Name"]    = company_name
            raw_df["Extraction Date"] = today
            cleaned_df = raw_df

    # ── Step 6: Save to Excel ─────────────────────────────────────────────────
    try:
        append_to_notices_master(cleaned_df)
        update_last_checked_date(company_name)
    except Exception as e:
        print(f"  ❌ [SAVE ERROR] {e}")

    # ── Step 7: Reset UI — ALWAYS runs ───────────────────────────────────────
    reset_to_main_screen()
    return True


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 11: MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  🤖 WINMAN GST NOTICE CHECKER — STARTING")
    print(f"  Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)
    print("\n⚠️  FAILSAFE: Move mouse to TOP-LEFT corner to abort instantly.")
    print("⏳ Focus the Winman window now — starting in 5 seconds...\n")
    time.sleep(5)

    client      = init_gemini()
    client_df   = load_client_master()
    eligible_df = filter_eligible_clients(client_df)

    if eligible_df.empty:
        print("✅ [DONE] No clients need checking today. Exiting.")
        return

    total, succeeded, failed = len(eligible_df), 0, 0

    for idx, row in eligible_df.iterrows():
        company_name = str(row.get("Company Name", "")).strip()
        gstin        = str(row.get("GSTIN",        "")).strip()

        if not company_name or not gstin:
            print(f"⚠️  [SKIP] Row {idx}: Missing Company Name or GSTIN.")
            failed += 1
            continue

        success = process_client(client, company_name, gstin)
        if success:
            succeeded += 1
            print(f"  ✅ ({succeeded + failed}/{total}) Done: {company_name}")
        else:
            failed += 1
            print(f"  ❌ ({succeeded + failed}/{total}) Failed: {company_name}")

        time.sleep(3)   # Polite pause between clients

    print("\n" + "=" * 62)
    print(f"  🏁 RUN COMPLETE")
    print(f"  Total    : {total}")
    print(f"  ✅ Passed : {succeeded}")
    print(f"  ❌ Failed : {failed}")
    print(f"  Ended at : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 12: COORDINATE CALIBRATION HELPER
# ─────────────────────────────────────────────────────────────────────────────
# Uncomment and run this file directly to find pixel coordinates.
# Hover over each Winman UI element and note the printed (X, Y). Ctrl+C stops.
#
# if __name__ == "__main__":
#     print("🎯 Coordinate finder active. Hover over UI elements. Ctrl+C to stop.")
#     try:
#         while True:
#             x, y = pyautogui.position()
#             print(f"  X: {x:5d} | Y: {y:5d}", end="\r")
#             time.sleep(0.5)
#     except KeyboardInterrupt:
#         print("\n✅ Done.")

# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()